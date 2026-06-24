#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Analise multiano (2023-2025, atual governo) da rede de coparticipacao organizacional
nas audiencias, debates e comissoes gerais da Camara dos Deputados.

Une todos os arquivos eventos-*.json de um diretorio, deduplica por id, bina por
ano-calendario (dataHoraInicio) e, para os anos COMPLETOS, calcula por ano as
mesmas metricas do pipeline de ano unico (cobertura, rede, composicao setorial,
centralizacao de Freeman, homofilia por permutacao) e, entre anos consecutivos, a
RETENCAO de diades e a PERSISTENCIA das organizacoes. 2026 entra apenas como linha
descritiva parcial, fora da comparacao e da retencao.

A extracao das organizacoes da pauta e o classificador setorial sao reutilizados,
sem alteracao, do modulo extracao_camara.py (fatorado da versao validada).

Uso:
  python3 analise_multiano.py [diretorio_dos_dados]
  (ou variavel de ambiente DADOS_DIR; saida em ./saida_multiano ou SAIDA_DIR)
"""
import json, os, sys, glob, statistics, random
from collections import Counter, defaultdict
import networkx as nx
import extracao_camara as ec

random.seed(42)
DADOS = sys.argv[1] if len(sys.argv) > 1 else os.environ.get("DADOS_DIR", "dados")
OUT = os.environ.get("SAIDA_DIR", "saida_multiano")
os.makedirs(OUT, exist_ok=True)
ANOS_COMPLETOS = ["2023", "2024", "2025"]
ANO_PARCIAL = "2026"

# Crosswalk de equivalencias para resolucao de identidade entre anos (ponto de
# extensao). Dentro de um mesmo governo a estrutura ministerial e' estavel, de modo
# que por padrao este dicionario fica vazio; variantes de grafia ja sao tratadas por
# ec.norm. chave: forma normalizada variante -> forma normalizada canonica.
CROSSWALK = {}


def canon(display):
    k = ec.norm(display)
    return CROSSWALK.get(k, k)


def is_alvo(e):
    cod = str(e.get("codTipoEvento") or e.get("idTipoEvento") or "")
    tt = (e.get("descricaoTipo") or "").lower()
    return cod in ec.TIPOS_ALVO or any(
        k in tt for k in ["audiência", "audiencia", "debate", "comissão geral", "comissao geral"])


def permuta(ev_orgs, labmap, nperm=1000, seed=42):
    rnd = random.Random(seed)
    ev_struct = [[labmap[o] for o in v if labmap[o] != "Indefinido"] for v in ev_orgs.values()]
    elig = [i for i, s in enumerate(ev_struct) if len(s) >= 2]
    if not elig:
        return dict(obs=0.0, mu=0.0, sd=0.0, z=float("nan"), p=1.0, nelig=0)
    obs = 100.0 * sum(1 for i in elig if len(set(ev_struct[i])) >= 2) / len(elig)
    stubs = [s for secs in ev_struct for s in secs]
    sizes = [len(secs) for secs in ev_struct]
    null = []
    for _ in range(nperm):
        rnd.shuffle(stubs); pos = 0; multi = tot = 0
        for sz in sizes:
            if sz >= 2:
                tot += 1
                if len(set(stubs[pos:pos + sz])) >= 2: multi += 1
            pos += sz
        null.append(100.0 * multi / tot if tot else 0.0)
    mu = statistics.mean(null); sd = statistics.pstdev(null)
    z = (obs - mu) / sd if sd else float("nan")
    ex = sum(1 for s in null if abs(s - mu) >= abs(obs - mu)); p = (ex + 1) / (nperm + 1)
    return dict(obs=round(obs, 2), mu=round(mu, 2), sd=round(sd, 2),
                z=round(z, 2), p=round(p, 4), nelig=len(elig))


def robustez_ruido(ev_orgs, labmap, mu, sd, trials=400, seed=123,
                   niveis=(5, 10, 20, 30, 40, 50, 55, 60, 70, 80, 90, 100)):
    """Sensibilidade do teste de homofilia a erro de rotulagem. Embaralha os rotulos
    setoriais de uma fracao p das organizacoes classificadas (preserva quem e Indefinido,
    os tamanhos por evento e a distribuicao global de setores; logo mu e sd do nulo
    permanecem validos) e recalcula z = (obs_corrompido - mu)/sd contra o mesmo nulo de
    margens fixas do teste principal."""
    classific = [o for o in labmap if labmap[o] != "Indefinido"]
    base = {o: labmap[o] for o in classific}
    rnd = random.Random(seed)

    def obsc(lab):
        es = [[lab[o] for o in v if lab[o] != "Indefinido"] for v in ev_orgs.values()]
        el = [s for s in es if len(s) >= 2]
        return 100.0 * sum(1 for s in el if len(set(s)) >= 2) / len(el) if el else 0.0

    res = []
    for pct in niveis:
        p = pct / 100.0
        zs = []
        for _ in range(trials):
            lab = dict(labmap)
            k = int(round(p * len(classific)))
            ch = rnd.sample(classific, k) if k else []
            vals = [base[o] for o in ch]
            rnd.shuffle(vals)
            for o, nv in zip(ch, vals):
                lab[o] = nv
            zs.append((obsc(lab) - mu) / sd if sd else float("nan"))
        zm = round(statistics.mean(zs), 2)
        res.append({"ruido_pct": pct, "z_medio": zm, "significativo_5pct": zm < -1.96})
    signif = [r["ruido_pct"] for r in res if r["significativo_5pct"]]
    return {"trials": trials, "seed": seed,
            "limiar_significancia_5pct_pct": max(signif) if signif else 0,
            "niveis": res}


def extrai_ano(eventos):
    ev_orgs, canon_display, org_freq = {}, {}, Counter()
    alvo = [e for e in eventos if is_alvo(e)]
    for e in alvo:
        ks = []
        for o in ec.extrair(e.get("descricao")):
            k = canon(o); canon_display.setdefault(k, o); ks.append(k)
        ks = sorted(set(ks))
        if ks:
            ev_orgs[e.get("id")] = ks
            for k in ks: org_freq[k] += 1
    return len(alvo), ev_orgs, canon_display, org_freq


def metricas(n_alvo, ev_orgs, canon_display):
    co = Counter()
    for v in ev_orgs.values():
        for i in range(len(v)):
            for j in range(i + 1, len(v)):
                co[(v[i], v[j])] += 1
    G = nx.Graph()
    for (a, b), w in co.items(): G.add_edge(a, b, weight=w)
    nodes = set()
    for v in ev_orgs.values(): nodes.update(v)
    G.add_nodes_from(nodes)
    n_nos = G.number_of_nodes(); n_ar = G.number_of_edges()
    dens = nx.density(G); comps = list(nx.connected_components(G))
    gig = max(comps, key=len) if comps else set()
    Gg = G.subgraph(gig).copy(); Ng = Gg.number_of_nodes()
    deg = nx.degree_centrality(Gg); bet = nx.betweenness_centrality(Gg, normalized=True)
    if Ng > 2:
        dmax = max(deg.values()); bmax = max(bet.values())
        Cd = sum(dmax - x for x in deg.values()) / (Ng - 2)
        Cb = sum(bmax - x for x in bet.values()) / (Ng - 1)
    else:
        Cd = Cb = float("nan")
    topbet = [(canon_display[k], round(v, 5)) for k, v in sorted(bet.items(), key=lambda x: -x[1])[:10]]
    labmap = {k: ec.setor_new(canon_display[k]) for k in canon_display}
    comp = Counter(labmap.values())
    n_multiorg = sum(1 for v in ev_orgs.values() if len(v) >= 2)
    n_mss = sum(1 for v in ev_orgs.values()
                if len({labmap[o] for o in v if labmap[o] != "Indefinido"}) >= 2)
    pesos = Counter(w for w in co.values())
    edge_set = {frozenset((a, b)) for (a, b) in co}
    org_set = set(nodes)
    perm_new = permuta(ev_orgs, labmap)
    perm_old = permuta(ev_orgs, {k: ec.setor_old(canon_display[k]) for k in canon_display})
    n_com_org = len(ev_orgs)
    return dict(
        n_alvo=n_alvo, n_com_org=n_com_org,
        cobertura=round(100 * n_com_org / n_alvo, 1) if n_alvo else 0,
        n_nos=n_nos, n_arestas=n_ar, densidade=round(dens, 4),
        n_componentes=len(comps), n_gigante=Ng,
        pct_gigante=round(100 * Ng / n_nos, 1) if n_nos else 0,
        n_multiorg=n_multiorg, pct_multiorg=round(100 * n_multiorg / n_alvo, 1) if n_alvo else 0,
        n_multissetor=n_mss, pct_multissetor=round(100 * n_mss / n_alvo, 1) if n_alvo else 0,
        comp={s: comp.get(s, 0) for s in ["Estado", "Mercado", "Terceiro setor", "Indefinido"]},
        pct_indef=round(100 * comp.get("Indefinido", 0) / n_nos, 1) if n_nos else 0,
        recorrencia=dict(sorted(pesos.items())),
        pct_diades_pontuais=round(100 * pesos.get(1, 0) / n_ar, 1) if n_ar else 0.0,
        C_grau=round(100 * Cd, 1), C_interm=round(100 * Cb, 1),
        top_intermediacao=topbet, perm_new=perm_new, perm_old=perm_old,
        _edge_set=edge_set, _org_set=org_set,
    )


def retencao(a, b, key):
    A, B = a[key], b[key]
    inter = len(A & B); uni = len(A | B)
    ret = round(100 * inter / len(A), 1) if A else 0.0
    return dict(n_t=len(A), n_t1=len(B), intersecao=inter,
                jaccard=round(inter / uni, 4) if uni else 0.0,
                retencao_de_t=ret, nao_retencao_de_t=round(100 - ret, 1))


# ---------- carga: pool + dedup + binagem ----------
eventos_por_ano = defaultdict(list); vistos = set()
arquivos = sorted(glob.glob(os.path.join(DADOS, "eventos-*.json")))
if not arquivos:
    print(f"Nenhum eventos-*.json em {DADOS}"); sys.exit(1)
n_total = 0
for f in arquivos:
    d = json.load(open(f, encoding="utf-8"))
    for e in (d.get("dados", d) if isinstance(d, dict) else d):
        i = e.get("id")
        if i in vistos: continue
        vistos.add(i); n_total += 1
        dh = e.get("dataHoraInicio") or ""
        if len(dh) >= 4: eventos_por_ano[dh[:4]].append(e)
print(f"Arquivos: {len(arquivos)} | eventos unicos (pos-dedup): {n_total}")
print(f"Eventos por ano-calendario: " + ", ".join(f"{a}={len(eventos_por_ano[a])}" for a in sorted(eventos_por_ano)))

# ---------- metricas por ano ----------
M = {}
anos_proc = ANOS_COMPLETOS + ([ANO_PARCIAL] if eventos_por_ano.get(ANO_PARCIAL) else [])
for a in anos_proc:
    n_alvo, ev_orgs, cd, freq = extrai_ano(eventos_por_ano.get(a, []))
    M[a] = metricas(n_alvo, ev_orgs, cd)
    M[a]["_freq"] = freq; M[a]["_canon_display"] = cd; M[a]["_ev_orgs"] = ev_orgs

print("\n=== RESUMO POR ANO ===")
print(f'{"ano":>5} {"alvo":>5} {"cob%":>5} {"nos":>5} {"arest":>6} {"Est":>4} {"Mer":>4} {"Ter":>4} {"Ind%":>5} {"mss%":>5} {"zNew":>6} {"Cgrau":>5} {"Cintr":>5}')
for a in anos_proc:
    m = M[a]; c = m["comp"]
    tag = "*" if a == ANO_PARCIAL else " "
    print(f'{a+tag:>5} {m["n_alvo"]:>5} {m["cobertura"]:>5} {m["n_nos"]:>5} {m["n_arestas"]:>6} '
          f'{c["Estado"]:>4} {c["Mercado"]:>4} {c["Terceiro setor"]:>4} {m["pct_indef"]:>5} '
          f'{m["pct_multissetor"]:>5} {m["perm_new"]["z"]:>6} {m["C_grau"]:>5.1f} {m["C_interm"]:>5.1f}')

# ---------- retencao entre anos completos ----------
RET = {"diadica": [], "organizacoes": []}
for x, y in zip(ANOS_COMPLETOS, ANOS_COMPLETOS[1:]):
    rd = retencao(M[x], M[y], "_edge_set"); rd["par"] = f"{x}-{y}"
    ro = retencao(M[x], M[y], "_org_set"); ro["par"] = f"{x}-{y}"
    RET["diadica"].append(rd); RET["organizacoes"].append(ro)
print("\n=== RETENCAO DIADICA (anos completos) ===")
for r in RET["diadica"]:
    print(f'  {r["par"]}: |E_t|={r["n_t"]} |E_t+1|={r["n_t1"]} inter={r["intersecao"]} '
          f'Jaccard={r["jaccard"]} retencao_de_t={r["retencao_de_t"]}%')

# ---------- persistencia de organizacoes (anos completos) ----------
pres = defaultdict(set); disp = {}
for a in ANOS_COMPLETOS:
    for k in M[a]["_org_set"]:
        pres[k].add(a); disp.setdefault(k, M[a]["_canon_display"][k])
por_nanos = Counter(len(v) for v in pres.values())
persistentes = sorted([disp[k] for k, v in pres.items() if len(v) == len(ANOS_COMPLETOS)])
# brokers persistentes: top-15 intermediacao em >=2 anos completos
topk = {a: {o for o, _ in M[a]["top_intermediacao"][:15]} for a in ANOS_COMPLETOS}
cont_broker = Counter()
for a in ANOS_COMPLETOS:
    for o in topk[a]: cont_broker[o] += 1
brokers_persist = sorted([o for o, c in cont_broker.items() if c >= 2], key=lambda o: -cont_broker[o])
print("\n=== PERSISTENCIA ===")
print(f'  organizacoes por nº de anos presentes (de {len(ANOS_COMPLETOS)}): ' +
      ", ".join(f"{n}ano(s):{por_nanos[n]}" for n in sorted(por_nanos)))
print(f'  presentes nos 3 anos: {len(persistentes)} | brokers em >=2 anos: {len(brokers_persist)}')
print(f'  brokers persistentes: {brokers_persist[:12]}')

# ---------- robustez do teste de homofilia a ruido de rotulagem (ano mais recente) ----------
if "2025" in M:
    cd25 = M["2025"]["_canon_display"]
    ev25 = M["2025"]["_ev_orgs"]
    lab25 = {k: ec.setor_new(cd25[k]) for k in cd25}
    pn25 = M["2025"]["perm_new"]
    M["2025"]["robustez_ruido"] = robustez_ruido(ev25, lab25, pn25["mu"], pn25["sd"])
    rr = M["2025"]["robustez_ruido"]
    print(f'\n=== ROBUSTEZ A RUIDO DE ROTULAGEM (2025) ===')
    print(f'  significancia a 5% preservada ate {rr["limiar_significancia_5pct_pct"]}% de rotulos perturbados')
    print('  ' + " | ".join(f'{n["ruido_pct"]}%:z={n["z_medio"]}' for n in rr["niveis"]))

# ---------- JSON ----------
def limpa(m):
    return {k: v for k, v in m.items() if not k.startswith("_")}
out = dict(
    anos_completos=ANOS_COMPLETOS, ano_parcial=ANO_PARCIAL,
    eventos_unicos=n_total,
    eventos_por_ano={a: len(eventos_por_ano[a]) for a in sorted(eventos_por_ano)},
    por_ano={a: limpa(M[a]) for a in anos_proc},
    retencao=RET,
    persistencia=dict(por_n_anos=dict(por_nanos),
                      total_orgs_distintas=sum(por_nanos.values()),
                      n_presentes_3anos=len(persistentes),
                      presentes_3anos=persistentes,
                      brokers_persistentes=[{"organizacao": o, "anos_no_top15": cont_broker[o]}
                                            for o in brokers_persist]),
    crosswalk_aplicado=CROSSWALK,
)
json.dump(out, open(os.path.join(OUT, "resultados_multiano.json"), "w", encoding="utf-8"),
          ensure_ascii=False, indent=2)
print(f"\nresultados_multiano.json salvo em {OUT}/")

# ---------- XLSX ----------
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

FONT = "Arial"
hdr_font = Font(name=FONT, bold=True, color="FFFFFF")
hdr_fill = PatternFill("solid", start_color="305496")
cell_font = Font(name=FONT)
center = Alignment(horizontal="center")


def escreve(ws, headers, rows, larguras=None):
    ws.append(headers)
    for c in ws[1]:
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = center
    for r in rows:
        ws.append(r)
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.font = cell_font
    if larguras:
        for i, w in enumerate(larguras, 1):
            ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"


wb = Workbook()
ws = wb.active; ws.title = "Resumo_por_ano"
escreve(ws,
        ["Ano", "Eventos-alvo", "Com >=1 org", "Cobertura %", "Nós", "Arestas", "Densidade",
         "Componentes", "Gigante", "Gigante %", "Multiorg", "Multiorg %", "Multissetor",
         "Multissetor %", "Indefinido %", "z homofilia", "C. grau %", "C. interm. %", "Díades pontuais %"],
        [[a + (" (parcial)" if a == ANO_PARCIAL else ""), M[a]["n_alvo"], M[a]["n_com_org"],
          M[a]["cobertura"], M[a]["n_nos"], M[a]["n_arestas"], M[a]["densidade"],
          M[a]["n_componentes"], M[a]["n_gigante"], M[a]["pct_gigante"], M[a]["n_multiorg"],
          M[a]["pct_multiorg"], M[a]["n_multissetor"], M[a]["pct_multissetor"], M[a]["pct_indef"],
          M[a]["perm_new"]["z"], M[a]["C_grau"], M[a]["C_interm"], M[a]["pct_diades_pontuais"]]
         for a in anos_proc],
        [16, 12, 11, 11, 8, 8, 10, 12, 8, 9, 9, 10, 11, 13, 12, 11, 10, 12, 16])

ws = wb.create_sheet("Composicao_por_ano")
escreve(ws, ["Ano", "Estado", "Mercado", "Terceiro setor", "Indefinido", "Total nós"],
        [[a + (" (parcial)" if a == ANO_PARCIAL else ""), M[a]["comp"]["Estado"], M[a]["comp"]["Mercado"],
          M[a]["comp"]["Terceiro setor"], M[a]["comp"]["Indefinido"], M[a]["n_nos"]] for a in anos_proc],
        [16, 9, 9, 14, 11, 10])

ws = wb.create_sheet("Homofilia_permutacao")
hrows = []
for a in anos_proc:
    tag = " (parcial)" if a == ANO_PARCIAL else ""
    for rot, key in [("gazetteer (nova)", "perm_new"), ("heurística (antiga)", "perm_old")]:
        pp = M[a][key]
        hrows.append([a + tag, rot, pp["obs"], pp["mu"], pp["sd"], pp["z"], pp["p"], pp["nelig"]])
escreve(ws, ["Ano", "Rotulagem", "Observado %", "Nulo (média) %", "Desvio %", "z", "p", "Eventos elegíveis"],
        hrows, [16, 20, 12, 15, 10, 8, 9, 16])

ws = wb.create_sheet("Top_intermediacao")
rows = []
for a in anos_proc:
    for rank, (org, val) in enumerate(M[a]["top_intermediacao"], 1):
        rows.append([a + (" (parcial)" if a == ANO_PARCIAL else ""), rank, org, val])
escreve(ws, ["Ano", "Rank", "Organização", "Intermediação (norm.)"], rows, [16, 6, 60, 18])

ws = wb.create_sheet("Retencao_diadica")
escreve(ws, ["Par de anos", "Díades em t", "Díades em t+1", "Interseção", "Jaccard", "Retenção de t %", "Não recorrência %"],
        [[r["par"], r["n_t"], r["n_t1"], r["intersecao"], r["jaccard"], r["retencao_de_t"], r["nao_retencao_de_t"]]
         for r in RET["diadica"]], [14, 12, 14, 11, 9, 15, 16])

ws = wb.create_sheet("Retencao_organizacoes")
escreve(ws, ["Par de anos", "Orgs em t", "Orgs em t+1", "Interseção", "Jaccard", "Retenção de t %", "Não recorrência %"],
        [[r["par"], r["n_t"], r["n_t1"], r["intersecao"], r["jaccard"], r["retencao_de_t"], r["nao_retencao_de_t"]]
         for r in RET["organizacoes"]], [14, 11, 13, 11, 9, 15, 16])

ws = wb.create_sheet("Persistencia_orgs")
escreve(ws, ["Nº de anos presente (de 3)", "Nº de organizações"],
        [[n, por_nanos[n]] for n in sorted(por_nanos, reverse=True)], [26, 20])
ws.append(["Total de organizações distintas (triênio)", sum(por_nanos.values())])
ws["A" + str(ws.max_row)].font = Font(name=FONT, bold=True)
ws.append([]); ws.append(["Organizações presentes nos 3 anos:", len(persistentes)])
ws["A" + str(ws.max_row)].font = Font(name=FONT, bold=True)
for org in persistentes:
    ws.append([org])
    ws["A" + str(ws.max_row)].font = cell_font

ws = wb.create_sheet("Brokers_persistentes")
escreve(ws, ["Organização", "Anos no top-15 de intermediação (de 3)"],
        [[o, cont_broker[o]] for o in brokers_persist], [60, 34])

ws = wb.create_sheet("Recorrencia_pesos")
rrows = []
for a in anos_proc:
    tag = " (parcial)" if a == ANO_PARCIAL else ""
    for peso, n in sorted(M[a]["recorrencia"].items(), key=lambda x: int(x[0])):
        rrows.append([a + tag, int(peso), n])
escreve(ws, ["Ano", "Peso (nº de coaparições)", "Nº de díades"], rrows, [16, 26, 16])

if "2025" in M and "robustez_ruido" in M["2025"]:
    rr = M["2025"]["robustez_ruido"]
    ws = wb.create_sheet("Robustez_ruido")
    escreve(ws, ["Ruído (% rótulos perturbados)", "z médio (vs nulo de 2025)", "Significativo a 5%"],
            [[n["ruido_pct"], n["z_medio"], "sim" if n["significativo_5pct"] else "não"]
             for n in rr["niveis"]], [30, 26, 18])
    ws.append(["Limiar de significância a 5% (%)", rr["limiar_significancia_5pct_pct"]])
    ws["A" + str(ws.max_row)].font = Font(name=FONT, bold=True)
    ws.append(["Repetições por nível (trials)", rr["trials"]])
    ws.append(["Semente aleatória (seed)", rr["seed"]])

ws = wb.create_sheet("Metadados")
mrows = [["Eventos únicos (pós-dedup)", n_total],
         ["Anos completos", ", ".join(ANOS_COMPLETOS)],
         ["Ano parcial (rótulo configurado)", ANO_PARCIAL],
         ["Crosswalk aplicado", str(CROSSWALK)]]
for a in sorted(eventos_por_ano):
    mrows.append([f"Eventos no ano-calendário {a}", len(eventos_por_ano[a])])
escreve(ws, ["Campo", "Valor"], mrows, [40, 40])

wb.save(os.path.join(OUT, "resultados_multiano.xlsx"))
print(f"resultados_multiano.xlsx salvo em {OUT}/ ({len(wb.sheetnames)} abas)")
